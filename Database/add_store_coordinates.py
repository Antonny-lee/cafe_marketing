"""Adds lat/lng columns to stores (idempotent) and backfills them from the
lat/lng columns add_coordinates.py added to cafes.xlsx."""
import os

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

CAFES_XLSX = os.path.join(os.path.dirname(__file__), "..", "Web crawling", "Home", "cafes.xlsx")


def connect():
    return psycopg2.connect(
        host=os.environ["PG_POOL_HOST"],
        port=os.environ["PG_POOL_PORT"],
        dbname=os.environ["PG_POOL_DATABASE"],
        user=os.environ["PG_POOL_USER"],
        password=os.environ["PG_POOL_PASSWORD"],
    )


def main():
    conn = connect()
    cur = conn.cursor()

    # create_tables_postgres.sql already defines lat/lng, but IF NOT EXISTS keeps
    # this script safe to run standalone too.
    cur.execute("ALTER TABLE stores ADD COLUMN IF NOT EXISTS lat NUMERIC(9,6)")
    cur.execute("ALTER TABLE stores ADD COLUMN IF NOT EXISTS lng NUMERIC(9,6)")
    print("lat/lng 컬럼 확인 완료.")

    wb = load_workbook(CAFES_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    lat_idx, lng_idx = headers.index("lat"), headers.index("lng")

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        store_id, lat, lng = row[0], row[lat_idx], row[lng_idx]
        if lat and lng:
            data.append((float(lat), float(lng), store_id))

    cur.executemany("UPDATE stores SET lat = %s, lng = %s WHERE store_id = %s", data)
    conn.commit()
    print(f"좌표 {len(data)}건 반영 완료.")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()

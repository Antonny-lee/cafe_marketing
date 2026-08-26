"""Incrementally reloads ONE store's crawled data (reviews + review tags) into
Oracle from the xlsx files, without touching any other store's rows.

Assumes Review/Review.py and Review/Review_category.py have already been
re-run (scoped to this store) so the xlsx files hold this store's latest data.

Usage:
    python refresh_store.py --store-id=S014
"""
import json
import os
import re
import sys
from datetime import date

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

BASE = os.path.join(os.path.dirname(__file__), "..", "Web crawling")
REVIEW_DIR = os.path.join(BASE, "Review", "reviews_by_store")
REVIEW_CATEGORY_DIR = os.path.join(BASE, "Review", "tags_by_store")


def parse_target_store_id():
    for a in sys.argv[1:]:
        if a.startswith("--store-id="):
            return a.split("=", 1)[1]
    return None


def connect():
    return psycopg2.connect(
        host=os.environ["PG_POOL_HOST"],
        port=os.environ["PG_POOL_PORT"],
        dbname=os.environ["PG_POOL_DATABASE"],
        user=os.environ["PG_POOL_USER"],
        password=os.environ["PG_POOL_PASSWORD"],
    )


def rows(path, sheet=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    it = ws.iter_rows(values_only=True)
    headers = next(it)
    for r in it:
        if r[0] is None:
            continue
        yield dict(zip(headers, r))


def parse_review_date(text):
    if not text:
        return None
    parts = [p for p in str(text).strip().split(".") if p]
    try:
        if len(parts) == 4:
            yy, m, d, _wd = parts
            year = 2000 + int(yy)
        elif len(parts) == 3:
            m, d, _wd = parts
            year = date.today().year
        else:
            return None
        return date(year, int(m), int(d))
    except ValueError:
        return None


def parse_visit_count(text):
    if not text:
        return None
    m = re.search(r"(\d+)", str(text))
    return int(m.group(1)) if m else None


def refresh_reviews(cur, store_id):
    cur.execute("SELECT review_id FROM reviews WHERE store_id = %s", [store_id])
    existing_ids = {r[0] for r in cur.fetchall()}

    review_path = os.path.join(REVIEW_DIR, f"{store_id}.xlsx")
    if not os.path.exists(review_path):
        return 0

    data = []
    for r in rows(review_path):
        if r["store_id"] != store_id or r["review_id"] in existing_ids:
            continue
        d = parse_review_date(r["review_date"])
        data.append((
            r["review_id"], r["store_id"], r["reviewer_id"],
            float(r["rating"]) if r["rating"] not in (None, "") else None,
            r["visit_time"], r["wait_time"], r["tags"], r["review_text"],
            r["review_date"], d, r["visit_count"], parse_visit_count(r["visit_count"]),
        ))

    if data:
        cur.executemany(
            "INSERT INTO reviews (review_id, store_id, reviewer_id, rating, visit_time, wait_time, "
            "tags, review_text, review_date_text, review_date, visit_count_text, visit_count) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
            data,
        )
    return len(data)


def refresh_tags(cur, store_id):
    best = {}
    tag_path = os.path.join(REVIEW_CATEGORY_DIR, f"{store_id}.xlsx")
    if os.path.exists(tag_path):
        for r in rows(tag_path):
            if r["store_id"] != store_id:
                continue
            key = r["tag_text"]
            count = r["mention_count"] or 0
            if key in best and count <= (best[key][2] or 0):
                continue
            best[key] = (r["store_id"], r["tag_text"], count, r["tag_category"], r["store_total_participants"])
    data = list(best.values())

    cur.execute("DELETE FROM review_category_tags WHERE store_id = %s", [store_id])
    if data:
        cur.executemany(
            "INSERT INTO review_category_tags (store_id, tag_text, mention_count, tag_category, "
            "store_total_participants) VALUES (%s, %s, %s, %s, %s)",
            data,
        )
    return len(data)


def main():
    store_id = parse_target_store_id()
    if not store_id:
        print("사용법: python refresh_store.py --store-id=S014")
        sys.exit(1)

    conn = connect()
    cur = conn.cursor()
    try:
        new_review_count = refresh_reviews(cur, store_id)
        print(f"reviews: 신규 {new_review_count}건 추가")

        tag_count = refresh_tags(cur, store_id)
        print(f"review_category_tags: {tag_count}건으로 갱신")

        conn.commit()
        print("RESULT_JSON:" + json.dumps({"newReviews": new_review_count, "tagRows": tag_count}))
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()

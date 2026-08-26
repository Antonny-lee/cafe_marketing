"""Loads every crawled Excel file + report.docx into the Postgres (Supabase)
schema. The schema itself is managed by create_tables_postgres.sql -- this
script only clears and reloads the 9 tables it owns, it doesn't create them.

Setup:
    1. pip install psycopg2-binary python-dotenv openpyxl
    2. Database/.env has PG_POOL_HOST/PORT/DATABASE/USER/PASSWORD filled in
    3. Run create_tables_postgres.sql once (Supabase SQL Editor or psql)
    4. python load_data.py
"""
import os
import re
import zipfile
from datetime import date, datetime

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

BASE = os.path.join(os.path.dirname(__file__), "..", "Web crawling")
CAFES_XLSX = os.path.join(BASE, "Home", "cafes.xlsx")
MENU_XLSX = os.path.join(BASE, "Menu", "menu.xlsx")
REVIEW_DIR = os.path.join(BASE, "Review", "reviews_by_store")
REVIEW_CATEGORY_DIR = os.path.join(BASE, "Review", "tags_by_store")
INFO_XLSX = os.path.join(BASE, "Info", "Info.xlsx")
AI_BRIEFING_XLSX = os.path.join(BASE, "AIBriefing", "AIBriefing.xlsx")
REPORT_DOCX = os.path.join(os.path.dirname(__file__), "..", "report.docx")

# Tables this script owns, in child-to-parent order so DELETE respects FKs.
TABLES_TO_CLEAR = [
    "ai_briefing", "store_info_items", "store_intro", "review_category_tags",
    "reviews", "menu", "market_report_metric", "market_report", "stores",
]


def connect():
    return psycopg2.connect(
        host=os.environ["PG_POOL_HOST"],
        port=os.environ["PG_POOL_PORT"],
        dbname=os.environ["PG_POOL_DATABASE"],
        user=os.environ["PG_POOL_USER"],
        password=os.environ["PG_POOL_PASSWORD"],
    )


def clear_tables(cur):
    print("기존 데이터 비우는 중 (자식 테이블부터)...")
    for table in TABLES_TO_CLEAR:
        cur.execute(f"DELETE FROM {table}")
    print("완료.")


def rows(path, sheet=None, header_row=1):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    it = ws.iter_rows(values_only=True)
    headers = next(it)
    for r in it:
        if r[0] is None:
            continue
        yield dict(zip(headers, r))


def iter_store_files(directory):
    """Yields the path of every per-store xlsx file in a reviews_by_store/tags_by_store dir."""
    if not os.path.isdir(directory):
        return
    for filename in sorted(os.listdir(directory)):
        if filename.endswith(".xlsx"):
            yield os.path.join(directory, filename)


def load_stores(cur):
    print("stores 적재 중...")
    data = [
        (
            r["store_id"], r["가게이름"], r["위치"], r["간편위치"], r["영업시간"], str(r["naver_id"]),
            float(r["lat"]) if r.get("lat") else None, float(r["lng"]) if r.get("lng") else None,
        )
        for r in rows(CAFES_XLSX)
    ]
    cur.executemany(
        "INSERT INTO stores (store_id, name, address, subway_info, business_hours, naver_id, lat, lng) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
        data,
    )
    print(f"  {len(data)}건")


def to_price(value):
    if isinstance(value, (int, float)):
        return int(value), None
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip()), None
    return None, (str(value).strip() if value not in (None, "") else None)


def load_menu(cur):
    print("menu 적재 중...")
    data = []
    for r in rows(MENU_XLSX):
        price_krw, price_note = to_price(r["price_krw"])
        data.append((r["menu_id"], r["store_id"], r["menu_name"], price_krw, price_note))
    cur.executemany(
        "INSERT INTO menu (menu_id, store_id, menu_name, price_krw, price_note) VALUES (%s, %s, %s, %s, %s)",
        data,
    )
    print(f"  {len(data)}건")


def parse_review_date(text):
    if not text:
        return None
    parts = [p for p in str(text).strip().split(".") if p]
    try:
        if len(parts) == 4:
            yy, m, d, _wd = parts
            year = 2000 + int(yy)
        elif len(parts) == 3:
            m, d, _wd = parts
            year = date.today().year
        else:
            return None
        return date(year, int(m), int(d))
    except ValueError:
        return None


def parse_visit_count(text):
    if not text:
        return None
    m = re.search(r"(\d+)", str(text))
    return int(m.group(1)) if m else None


def load_reviews(cur):
    print("reviews 적재 중 (시간이 좀 걸립니다)...")
    data = []
    for path in iter_store_files(REVIEW_DIR):
        for r in rows(path):
            d = parse_review_date(r["review_date"])
            data.append((
                r["review_id"], r["store_id"], r["reviewer_id"],
                float(r["rating"]) if r["rating"] not in (None, "") else None,
                r["visit_time"], r["wait_time"], r["tags"], r["review_text"],
                r["review_date"], d, r["visit_count"], parse_visit_count(r["visit_count"]),
            ))
    cur.executemany(
        "INSERT INTO reviews (review_id, store_id, reviewer_id, rating, visit_time, wait_time, "
        "tags, review_text, review_date_text, review_date, visit_count_text, visit_count) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
        data,
    )
    print(f"  {len(data)}건")


def load_review_category(cur):
    print("review_category_tags 적재 중...")
    best = {}
    dup_count = 0
    for path in iter_store_files(REVIEW_CATEGORY_DIR):
        for r in rows(path):
            key = (r["store_id"], r["tag_text"])
            count = r["mention_count"] or 0
            if key in best:
                dup_count += 1
                if count <= (best[key][2] or 0):
                    continue
            best[key] = (r["store_id"], r["tag_text"], count, r["tag_category"], r["store_total_participants"])
    data = list(best.values())
    cur.executemany(
        "INSERT INTO review_category_tags (store_id, tag_text, mention_count, tag_category, "
        "store_total_participants) VALUES (%s, %s, %s, %s, %s)",
        data,
    )
    print(f"  {len(data)}건 (중복 {dup_count}건 중 mention_count 낮은 쪽 제거)")


def load_info(cur):
    print("store_intro / store_info_items 적재 중...")
    intro = [(r["store_id"], r["intro_text"]) for r in rows(INFO_XLSX, sheet="소개")]
    cur.executemany("INSERT INTO store_intro (store_id, intro_text) VALUES (%s, %s)", intro)

    item_count = 0
    for sheet in ["편의시설 및 서비스", "노키즈존", "반려동물 동반", "주차", "좌석.공간", "결제수단", "SNS"]:
        section_rows = [(r["store_id"], sheet, r["item_text"], r["detail"]) for r in rows(INFO_XLSX, sheet=sheet)]
        cur.executemany(
            "INSERT INTO store_info_items (store_id, section, item_text, detail) VALUES (%s, %s, %s, %s)",
            section_rows,
        )
        item_count += len(section_rows)
    print(f"  소개 {len(intro)}건, 나머지 섹션 {item_count}건")


def load_ai_briefing(cur):
    print("ai_briefing 적재 중...")
    data = [(r["store_id"], r["sentence"]) for r in rows(AI_BRIEFING_XLSX)]
    cur.executemany("INSERT INTO ai_briefing (store_id, sentence) VALUES (%s, %s)", data)
    print(f"  {len(data)}건")


def extract_docx_text(path):
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8")
    import html
    text = re.sub(r"<[^>]+>", "", xml)
    return html.unescape(text)


METRIC_PATTERNS = [
    ("점포수", r"점포수는\s*([\d,]+)\s*개\s*입니다", "개"),
    ("생존율(3년)", r"신생기업 생존율은\s*([\d.]+)%\s*입니다", "%"),
    ("평균영업기간", r"평균\s*영업기간은\s*([\d.]+)\s*년\s*입니다", "년"),
    ("개업수", r"개업수은\s*([\d,]+)\s*개\s*입니다", "개"),
    ("폐업수", r"페업수는\s*([\d,]+)\s*개\s*입니다", "개"),
    ("매출액(월평균)", r"점포당 월평균 매출액은\s*([\d,]+)\s*만원\s*입니다", "만원"),
    ("매출건수(월평균)", r"점포당 월평균 매출건수는\s*([\d,]+)\s*건\s*입니다", "건"),
    ("유동인구(일평균)", r"유동인구\s*수는\s*일평균\s*([\d,]+)\s*명", "명"),
    ("유동인구밀도", r"밀도는\s*([\d,]+)\s*명\s*/\s*ha\s*입니다", "명/ha"),
    ("주거인구", r"주거인구는\s*([\d,]+)\s*명", "명"),
    ("직장인구", r"직장인구\s*수는\s*([\d,]+)\s*명", "명"),
    ("가구세대수", r"가구세대\s*수는\s*([\d,]+)\s*가구\s*입니다", "가구"),
    ("임대시세(1층, 3.3㎡당)", r"1층\s*임대료가\s*3\.3㎡당\s*([\d,]+)\s*원입니다", "원"),
    ("소득수준", r"소득수준은\s*0?(\d+)\s*분위입니다", "분위"),
]


def load_market_report(cur):
    if not os.path.exists(REPORT_DOCX):
        print("report.docx 없음, 건너뜀")
        return
    print("market_report 적재 중...")
    text = extract_docx_text(REPORT_DOCX)

    loc_m = re.search(r"위치\s*([^업]+?)업종", text)
    industry_m = re.search(r"업종\s*([가-힣\-·]+)기분준기", text)
    quarter_m = re.search(r"(\d{4}년\s*\d분기)", text)
    date_m = re.search(r"(\d{4})년\s*(\d{2})월\s*(\d{2})일", text)

    location = loc_m.group(1).strip() if loc_m else None
    industry = industry_m.group(1).strip() if industry_m else None
    quarter = quarter_m.group(1) if quarter_m else None
    report_date = datetime(int(date_m.group(1)), int(date_m.group(2)), int(date_m.group(3))) if date_m else None

    cur.execute(
        "INSERT INTO market_report (location, industry, quarter, report_date, raw_text) "
        "VALUES (%s, %s, %s, %s, %s) RETURNING report_id",
        [location, industry, quarter, report_date, text],
    )
    report_id = cur.fetchone()[0]

    metrics = []
    for name, pattern, unit in METRIC_PATTERNS:
        m = re.search(pattern, text)
        if m:
            metrics.append((report_id, name, m.group(1).replace(",", ""), unit, None, None, None))
    if metrics:
        cur.executemany(
            "INSERT INTO market_report_metric (report_id, metric_name, value, unit, qoq_change, "
            "yoy_change, note) VALUES (%s, %s, %s, %s, %s, %s, %s)",
            metrics,
        )
    print(f"  리포트 1건 (위치={location}, 업종={industry}, 분기={quarter}), 지표 {len(metrics)}건 추출")


def main():
    conn = connect()
    cur = conn.cursor()
    try:
        clear_tables(cur)
        conn.commit()

        load_stores(cur)
        conn.commit()
        load_menu(cur)
        conn.commit()
        load_review_category(cur)
        conn.commit()
        load_info(cur)
        conn.commit()
        load_ai_briefing(cur)
        conn.commit()
        load_market_report(cur)
        conn.commit()
        load_reviews(cur)  # last: by far the largest table
        conn.commit()

        print("\n모든 데이터 적재 완료.")
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()

import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from scrapling.fetchers import Fetcher
from scrapling.parser import Selector

CAFES_XLSX = "cafes.xlsx"
OUTPUT_XLSX = "menu.xlsx"
MENU_URL_TMPL = "https://m.place.naver.com/place/{pid}/menu/list"
REQUEST_DELAY_SEC = 0.6


def load_stores(path: str):
    wb = load_workbook(path)
    ws = wb.active
    stores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        store_id, name, _loc, _subway, _hours, naver_id = row[:6]
        if not naver_id:
            print(f"  건너뜀 (naver_id 없음): {store_id} {name}")
            continue
        stores.append({"store_id": store_id, "name": name, "naver_id": str(naver_id)})
    return stores


def get_menu_items(pid: str):
    url = MENU_URL_TMPL.format(pid=pid)
    page = Fetcher.get(url, stealthy_headers=True)
    sel = Selector(page.body.decode("utf-8"))

    items = []
    for li in sel.css("li.E2jtL"):
        name = li.css(".lPzHi::text").get()
        if not name:
            continue
        price_num = li.css(".p2H02 em::text").get()
        price = None
        if price_num:
            cleaned = price_num.replace(",", "").strip()
            if cleaned.isdigit():
                price = int(cleaned)
        else:
            fallback = li.css(".p2H02::text").get()
            if fallback and fallback.strip():
                price = fallback.strip()
        items.append({"menu_name": name.strip(), "price_krw": price})
    return items


def save_to_excel(rows: list[dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "메뉴크롤링"
    headers = ["menu_id", "store_id", "store_name_ref", "menu_name", "price_krw"]
    ws.append(headers)

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for col_letter, width in zip("ABCDE", [10, 10, 24, 30, 12]):
        ws.column_dimensions[col_letter].width = width
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = wrap

    for attempt in range(1, 6):
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"  {path} 파일이 열려 있어 저장할 수 없습니다. 닫아주세요. ({attempt}/5, 5초 후 재시도)")
            time.sleep(5)
    wb.save(path)


def main():
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 3

    print("cafes.xlsx에서 매장 목록 읽는 중...")
    stores = load_stores(CAFES_XLSX)
    print(f"총 {len(stores)}개 매장. {limit}개 처리합니다.")

    rows = []
    menu_seq = 1
    for i, store in enumerate(stores[:limit], 1):
        print(f"[{i}/{limit}] {store['name']} ({store['naver_id']}) 메뉴 가져오는 중...")
        try:
            items = get_menu_items(store["naver_id"])
        except Exception as e:
            print(f"  실패: {e}")
            items = []

        if not items:
            print("  메뉴 없음")

        for item in items:
            rows.append({
                "menu_id": f"M{menu_seq:03d}",
                "store_id": store["store_id"],
                "store_name_ref": store["name"],
                "menu_name": item["menu_name"],
                "price_krw": item["price_krw"] if item["price_krw"] is not None else "",
            })
            menu_seq += 1

        time.sleep(REQUEST_DELAY_SEC)

    save_to_excel(rows, OUTPUT_XLSX)
    print(f"완료: {len(rows)}개 메뉴 저장 -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()

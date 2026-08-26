import os
import random
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from scrapling.fetchers import Fetcher

CAFES_XLSX = "../Home/cafes.xlsx"
OUTPUT_DIR = "tags_by_store"  # one xlsx per store, matching Review.py's reviews_by_store
DETAIL_URL_TMPL = "https://m.place.naver.com/place/{pid}/review/visitor"
REQUEST_DELAY_RANGE_SEC = (0.4, 1.0)  # randomized instead of a fixed 0.6s -- same anti-fingerprint reasoning as Review.py


def output_path_for(store_id: str) -> str:
    return os.path.join(OUTPUT_DIR, f"{store_id}.xlsx")

# Naver's review-keyword codes are a fixed, site-wide vocabulary (not per-store),
# so one lookup table covers every store. Grouped along the same lines as
# Naver's own "analysis.themes" categories, plus a few (interior/photo/space/
# convenience) that themes doesn't cover but the vote-keyword list does.
CODE_TO_CATEGORY = {
    "dessert_good": "맛", "coffee_good": "맛", "drink_good": "맛", "food_good": "맛",
    "taste_healthy": "맛", "tea_good": "맛", "bread_good": "맛", "fresh": "맛",
    "interior_cool": "인테리어",
    "kind": "서비스", "food_fast": "서비스",
    "special_menu": "메뉴", "menu_good": "메뉴", "types_various": "메뉴",
    "alcohol_var": "메뉴", "course_good": "메뉴",
    "talk_good": "분위기", "atmosphere_calm": "분위기", "cozy": "분위기",
    "music_good": "분위기", "concept_unique": "분위기", "drink_alone": "분위기",
    "together": "분위기", "stay_long": "분위기",
    "photo_good": "사진",
    "store_clean": "청결도", "toilet_clean": "청결도",
    "view_good": "전망", "outdoor_good": "전망",
    "study_good": "편의", "comfy": "편의", "parking_easy": "편의", "pet_good": "편의",
    "spacious": "공간",
    "price_cheap": "가격", "price_worthy": "가격",
    "amount": "음식량", "large": "음식량",
    "eat_alone": "목적", "kid_good": "목적", "special_day": "목적",
    "gift_good": "목적", "party_good": "목적",
    "dessert_good_bingsu": "맛", "dessert_good_icecream": "맛", "snack_good": "맛",
    "custom_good": "서비스", "packaging_clean": "서비스",
    "book_many": "편의", "game_various": "편의", "play_var": "편의",
    "live_show_good": "분위기",
    "room_nice": "인테리어",
}


def extract_apollo_state(html: str):
    marker = "window.__APOLLO_STATE__ = "
    idx = html.find(marker)
    if idx == -1:
        return None
    start = idx + len(marker)
    i = start
    depth = 0
    in_str = False
    esc = False
    while i < len(html):
        c = html[i]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
        else:
            if c == '"':
                in_str = True
            elif c == "{":
                depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    i += 1
                    break
        i += 1
    return json_loads_safe(html[start:i])


def json_loads_safe(text: str):
    import json
    return json.loads(text)


def load_stores(path: str):
    wb = load_workbook(path)
    ws = wb.active
    stores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        store_id, name, _loc, _subway, _hours, naver_id = row[:6]
        if not naver_id:
            continue
        stores.append({"store_id": store_id, "name": name, "naver_id": str(naver_id)})
    return stores


def get_category_rows(store: dict):
    url = DETAIL_URL_TMPL.format(pid=store["naver_id"])
    page = Fetcher.get(url, stealthy_headers=True)
    html = page.body.decode("utf-8")
    apollo = extract_apollo_state(html)
    if apollo is None:
        return []

    stats = apollo.get(f"VisitorReviewStatsResult:{store['naver_id']}")
    if not stats:
        return []

    voted = (stats.get("analysis") or {}).get("votedKeyword")
    if not voted:
        return []

    participants = voted.get("userCount") or 0
    rows = []
    for d in voted.get("details") or []:
        code = d.get("code", "")
        rows.append({
            "store_id": store["store_id"],
            "store_name_ref": store["name"],
            "tag_text": d.get("displayName", ""),
            "mention_count": d.get("count", 0),
            "tag_category": CODE_TO_CATEGORY.get(code, "기타"),
            "store_total_participants": participants,
            "_code": code,
        })
    return rows


HEADERS = ["store_id", "store_name_ref", "tag_text", "mention_count", "tag_category", "store_total_participants"]


def save_to_excel(rows: list[dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "카테고리태그"
    ws.append(HEADERS)

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])

    widths = [10, 22, 26, 14, 12, 20]
    for col_letter, width in zip("ABCDEF", widths):
        ws.column_dimensions[col_letter].width = width
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = wrap

    for attempt in range(1, 6):
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"  {path} 파일이 열려 있어 저장할 수 없습니다. 닫아주세요. ({attempt}/5, 5초 후 재시도)")
            time.sleep(5)
    wb.save(path)


def parse_target_store_id():
    for a in sys.argv[1:]:
        if a.startswith("--store-id="):
            return a.split("=", 1)[1]
    return None


def main():
    target_store_id = parse_target_store_id()

    print("cafes.xlsx에서 매장 목록 읽는 중...")
    stores = load_stores(CAFES_XLSX)

    if target_store_id:
        store_list = [s for s in stores if s["store_id"] == target_store_id]
        if not store_list:
            print(f"store_id={target_store_id}를 {CAFES_XLSX}에서 찾을 수 없습니다.")
            return
    else:
        limit = int(sys.argv[1]) if len(sys.argv) > 1 else 2
        store_list = stores[:limit]

    limit = len(store_list)
    print(f"총 {len(stores)}개 매장 중 {limit}개 처리합니다. (매장별로 {OUTPUT_DIR}/<store_id>.xlsx에 저장)")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    unknown_codes = set()
    total_rows = 0
    for i, store in enumerate(store_list, 1):
        print(f"[{i}/{limit}] {store['name']} ({store['naver_id']}) 카테고리 태그 가져오는 중...")
        try:
            rows = get_category_rows(store)
        except Exception as e:
            print(f"  실패: {e}")
            rows = []
        print(f"  -> {len(rows)}개 태그 수집")
        for r in rows:
            if r["tag_category"] == "기타":
                unknown_codes.add((r["_code"], r["tag_text"]))
            r.pop("_code", None)

        # 매장 하나당 파일 하나 = 항상 그 매장의 최신 스냅샷으로 통째로 덮어씀
        # (다른 매장 파일과 병합할 필요가 없어짐)
        save_to_excel(rows, output_path_for(store["store_id"]))
        total_rows += len(rows)
        time.sleep(random.uniform(*REQUEST_DELAY_RANGE_SEC))

    if unknown_codes:
        print(f"\n매핑 안 된 코드 {len(unknown_codes)}개 (tag_category='기타'로 처리됨):")
        for code, text in sorted(unknown_codes):
            print(f"  {code} ({text})")

    print(f"완료: 총 {total_rows}개 태그 저장 -> {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()

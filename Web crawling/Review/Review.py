import asyncio
import os
import random
import re
import sys
import time
from datetime import date, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from scrapling.fetchers import AsyncStealthySession
from scrapling.parser import Selector

CAFES_XLSX = "../Home/cafes.xlsx"
OUTPUT_DIR = "reviews_by_store"  # one xlsx per store -- keeps dedup lookups fast and lets
                                  # different stores' crawls never touch the same file
REVIEW_URL_TMPL = "https://pcmap.place.naver.com/restaurant/{pid}/review/visitor"
MONTHS_BACK_DAYS = 182  # ~6 months
MAX_MORE_CLICKS = 1000    # generous safety cap per store (not a real limit, just anti-infinite-loop)
CLICK_WAIT_MS = 450
CLICKS_PER_SESSION = 80   # restart the browser after this many "더보기" clicks within one store;
                          # long click runs in a single tab visibly degrade (CPU/memory) over time.
                          # 200 let extreme stores run long enough to peg the CPU for 1h+ before ever
                          # restarting; 60 wasted too much time re-walking already-seen reviews after
                          # every restart. 80 is a middle ground.
INTER_STORE_DELAY_RANGE_SEC = (6, 14)  # randomized pause between stores -- a fixed delay is itself
                          # a bot fingerprint (every request exactly N seconds apart); jittering keeps
                          # the average pacing that avoided Naver's 429 while looking less mechanical
MAX_EMPTY_PASSES = 2      # give up on a store after this many consecutive session restarts that
                          # found zero new reviews (a very active store's already-collected prefix
                          # can grow long enough that a full session's clicks never walk past it)


def output_path_for(store_id: str) -> str:
    return os.path.join(OUTPUT_DIR, f"{store_id}.xlsx")

CUTOFF = date.today() - timedelta(days=MONTHS_BACK_DAYS)


def load_stores(path: str):
    wb = load_workbook(path)
    ws = wb.active
    stores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        store_id, name, _loc, _subway, _hours, naver_id = row[:6]
        if not naver_id:
            print(f"  건너뜀 (naver_id 없음): {store_id} {name}")
            continue
        stores.append({"store_id": store_id, "name": name, "naver_id": str(naver_id)})
    return stores


def normalize_for_key(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def make_key(store_id: str, r: dict):
    # reviewer_id uniquely identifies the author, so (store, author, date) is
    # enough to distinguish reviews even when two people post blank-text
    # (photo-only) reviews on the same day. Fall back to the review text only
    # if we failed to extract a reviewer id.
    reviewer_id = r.get("reviewer_id") or ""
    if reviewer_id:
        return (store_id, reviewer_id, normalize_for_key(r["review_date"]))
    return (store_id, normalize_for_key(r["review_date"]), normalize_for_key(r["review_text"]))


def load_existing_reviews(path: str):
    """Returns (seen_keys: set of make_key(...) tuples, next_seq: dict[store_id -> int])."""
    seen = set()
    next_seq = {}
    if not os.path.exists(path):
        return seen, next_seq

    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        (review_id, store_id, _name, reviewer_id, _rating, _vt, _wt,
         _tags, review_text, review_date, _vc) = row
        if not store_id:
            continue
        seen.add(make_key(store_id, {
            "reviewer_id": reviewer_id,
            "review_date": review_date,
            "review_text": review_text,
        }))
        try:
            seq = int(str(review_id).rsplit("-", 1)[-1])
            next_seq[store_id] = max(next_seq.get(store_id, 0), seq)
        except (ValueError, AttributeError):
            pass
    return seen, next_seq


def parse_review_date(text):
    if not text:
        return None
    parts = [p for p in text.strip().split(".") if p]
    try:
        if len(parts) == 4:
            yy, m, d, _wd = parts
            year = 2000 + int(yy)
        elif len(parts) == 3:
            m, d, _wd = parts
            year = date.today().year
        else:
            return None
        return date(year, int(m), int(d))
    except ValueError:
        return None


def extract_review(li):
    rating_els = li.css(".pui__6abRMf")
    rating_text = rating_els[0].get_all_text(strip=True, separator="") if rating_els else ""
    m = re.search(r"별점([\d.]+)점", rating_text)
    rating = float(m.group(1)) if m else None

    visit_time = ""
    wait_time = ""
    etc_tags = []
    for tag in li.css(".pui__-0Ter1 .pui__V8F9nN"):
        em_text = (tag.css("em::text").get() or "").strip()
        has_icon = bool(tag.css("svg"))
        raw_text = tag.get_all_text(strip=True)
        if has_icon:
            visit_time = em_text.replace("에 방문", "").strip()
        elif "대기 시간" in raw_text:
            wait_time = em_text
        elif em_text:
            etc_tags.append(em_text)

    body_els = li.css(".pui__vn15t2 a")
    review_text = body_els[0].get_all_text(strip=True, separator="\n") if body_els else ""

    date_text = li.css(".Vk05k time::text").get() or ""
    gfu_spans = li.css(".Vk05k .pui__gfuUIT")
    visit_count_text = gfu_spans[1].get_all_text(strip=True) if len(gfu_spans) > 1 else ""

    profile_href = li.css('a[data-pui-click-code="profile"]::attr(href)').get() or ""
    uid_match = re.search(r"/my/([^/]+)/review", profile_href)
    reviewer_id = uid_match.group(1) if uid_match else ""

    return {
        "reviewer_id": reviewer_id,
        "rating": rating,
        "visit_time": visit_time,
        "wait_time": wait_time,
        "tags": ", ".join(etc_tags),
        "review_text": review_text,
        "review_date": date_text,
        "visit_count": visit_count_text,
        "parsed_date": parse_review_date(date_text),
    }


# Extracts outerHTML of every currently-loaded review card, then removes them
# from the DOM so the page never accumulates more than one batch's worth of
# review cards (each card carries several inline SVG icons; thousands of them
# on one page balloon the renderer's memory and can crash the tab).
JS_EXTRACT_AND_TRIM = """
() => {
    const lis = Array.from(document.querySelectorAll('li.place_apply_pui'));
    const html = lis.map(el => el.outerHTML);
    lis.forEach(el => el.remove());
    return html;
}
"""


def make_page_action(collected: list, disk_seen_keys: set, local_keys: set, store_id: str,
                      stop_flag: dict, continue_flag: dict):
    oldest_seen = {"date": None}

    async def process_current_batch(page):
        try:
            batch_html = await page.evaluate(JS_EXTRACT_AND_TRIM)
        except Exception:
            batch_html = []

        for li_html in batch_html:
            li = Selector(li_html).css("li.place_apply_pui")
            if not li:
                continue
            r = extract_review(li[0])
            if r["parsed_date"] and (oldest_seen["date"] is None or r["parsed_date"] < oldest_seen["date"]):
                oldest_seen["date"] = r["parsed_date"]
            key = make_key(store_id, r)
            if r["parsed_date"] and r["parsed_date"] < CUTOFF:
                # reached the 6-month boundary -- genuinely done with this store
                stop_flag["stop"] = True
            elif key in disk_seen_keys:
                # already saved to Review.xlsx in a previous run of this script
                stop_flag["stop"] = True
            elif key in local_keys:
                # already collected earlier in THIS run (a prior session pass for
                # this same store) -- a session restart always re-lands on the
                # most-recent reviews first, so we must skip back over this
                # overlap rather than treat it as "caught up"
                pass
            else:
                local_keys.add(key)
                collected.append(r)

        if oldest_seen["date"]:
            days_covered = (date.today() - oldest_seen["date"]).days
            pct = min(100, round(days_covered / MONTHS_BACK_DAYS * 100))
            print(f"    진행: 누적 {len(local_keys)}개 수집 | 6개월 기준 {pct}% "
                  f"(최고 오래된 날짜 {oldest_seen['date']})")

    async def action(page):
        await page.wait_for_timeout(1200)

        sort_btn = page.locator('a.place_btn_option', has_text="최신순").first
        if await sort_btn.count() > 0:
            try:
                await sort_btn.click(timeout=5000)
                await page.wait_for_timeout(1000)
            except Exception:
                pass

        # capture + trim the initial (pre-click) batch too
        await process_current_batch(page)

        clicks = 0
        for _ in range(MAX_MORE_CLICKS):
            if stop_flag.get("stop"):
                break
            if clicks >= CLICKS_PER_SESSION:
                # this store still has more to load; hand off to a fresh
                # browser session rather than let this tab keep degrading
                continue_flag["continue"] = True
                break

            more_btn = page.locator("a.fvwqf")
            if await more_btn.count() == 0:
                break

            try:
                await more_btn.first.click(timeout=30000)
            except Exception:
                break
            clicks += 1
            await page.wait_for_timeout(CLICK_WAIT_MS)

            await process_current_batch(page)

        return page

    return action


async def get_store_reviews_once(session, pid: str, disk_seen_keys: set, local_keys: set, store_id: str):
    """One browser-session pass. Returns (new_reviews, needs_more_sessions)."""
    url = REVIEW_URL_TMPL.format(pid=pid)

    for attempt in range(1, 5):
        collected: list = []
        stop_flag: dict = {}
        continue_flag: dict = {}
        page = await session.fetch(
            url,
            page_action=make_page_action(collected, disk_seen_keys, local_keys, store_id, stop_flag, continue_flag),
            network_idle=True,
            timeout=600000,
        )
        if getattr(page, "status", 200) == 429:
            wait_s = 30 * attempt
            print(f"    429(요청 과다) 감지, {wait_s}초 대기 후 재시도 ({attempt}/4)")
            await asyncio.sleep(wait_s)
            continue
        return collected, continue_flag.get("continue", False)

    print("    429가 계속돼서 이번 세션은 건너뜁니다")
    return [], False


async def get_store_reviews(pid: str, disk_seen_keys: set, store_id: str, on_pass=None):
    """Scrapes a store's reviews, opening a fresh browser session for every
    pass and again whenever a very active store needs more than one pass
    (see CLICKS_PER_SESSION) -- long-lived tabs visibly slow down and bloat
    in memory the more they click, so a clean restart is cheaper than trying
    to keep one tab alive indefinitely.

    Each restart always re-lands on the most-recent reviews, so `local_keys`
    tracks everything already collected for this store across passes; it is
    used to skip back over that overlap without mistaking it for "caught up
    to previously-saved data" (that's what disk_seen_keys is for).

    If `on_pass` is given, it's awaited with each pass's new reviews right
    after that pass finishes, so a caller can persist them immediately --
    a store can need many restarts, and without this, killing the process
    (or a crash) mid-store would discard everything collected so far."""
    all_reviews: list = []
    local_keys: set = set()
    needs_more = True
    passes = 0
    empty_passes_in_a_row = 0
    while needs_more:
        passes += 1
        async with AsyncStealthySession(headless=True, disable_resources=True) as session:
            new_reviews, needs_more = await get_store_reviews_once(session, pid, disk_seen_keys, local_keys, store_id)
        all_reviews.extend(new_reviews)
        if on_pass and new_reviews:
            await on_pass(new_reviews)
        if passes > 1:
            print(f"    (브라우저 재시작 {passes}회차, 누적 {len(all_reviews)}개)")

        if new_reviews:
            empty_passes_in_a_row = 0
        else:
            empty_passes_in_a_row += 1
            if empty_passes_in_a_row >= MAX_EMPTY_PASSES:
                # each restart re-walks everything already collected before it can
                # find anything new; once that walk-back alone eats a whole
                # session's click budget, no session will ever get further --
                # accept what we have rather than loop forever
                print(f"    {empty_passes_in_a_row}회 연속 신규 리뷰 없음 -- 이 매장은 여기서 정리하고 다음으로 넘어갑니다")
                break
    disk_seen_keys.update(local_keys)
    return all_reviews


HEADERS = ["review_id", "store_id", "store_name_ref", "reviewer_id", "rating", "visit_time",
           "wait_time", "tags", "review_text", "review_date", "visit_count"]


def save_to_excel(new_rows: list[dict], path: str):
    """Appends new_rows to the existing workbook at `path`, or creates it if missing."""
    wrap = Alignment(wrap_text=True, vertical="top")

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "리뷰크롤링"
        ws.append(HEADERS)
        widths = [12, 10, 20, 12, 8, 10, 12, 24, 45, 12, 14]
        for col_letter, width in zip("ABCDEFGHIJK", widths):
            ws.column_dimensions[col_letter].width = width

    for row in new_rows:
        ws.append([row.get(h, "") for h in HEADERS])
        for cell in ws[ws.max_row]:
            cell.alignment = wrap

    for attempt in range(1, 6):
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"  {path} 파일이 열려 있어 저장할 수 없습니다. 닫아주세요. ({attempt}/5, 5초 후 재시도)")
            time.sleep(5)
    wb.save(path)


def review_to_row(store: dict, r: dict, seq: int) -> dict:
    return {
        "review_id": f"{store['store_id']}-{seq:04d}",
        "store_id": store["store_id"],
        "store_name_ref": store["name"],
        "reviewer_id": r["reviewer_id"],
        "rating": r["rating"] if r["rating"] is not None else "",
        "visit_time": r["visit_time"],
        "wait_time": r["wait_time"],
        "tags": r["tags"],
        "review_text": r["review_text"],
        "review_date": r["review_date"],
        "visit_count": r["visit_count"],
    }


def format_eta(seconds: float) -> str:
    seconds = int(seconds)
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}시간 {m}분"
    if m:
        return f"{m}분 {s}초"
    return f"{s}초"


def parse_target_store_id():
    for a in sys.argv[1:]:
        if a.startswith("--store-id="):
            return a.split("=", 1)[1]
    return None


async def main():
    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    target_store_id = parse_target_store_id()

    print("cafes.xlsx에서 매장 목록 읽는 중...")
    stores = load_stores(CAFES_XLSX)

    if target_store_id:
        store_list = [s for s in stores if s["store_id"] == target_store_id]
        if not store_list:
            print(f"store_id={target_store_id}를 {CAFES_XLSX}에서 찾을 수 없습니다.")
            return
    else:
        limit = int(sys.argv[1]) if len(sys.argv) > 1 else 1
        store_list = stores[:limit]

    limit = len(store_list)
    print(f"총 {len(stores)}개 매장 중 {limit}개 처리합니다. (매장별로 {OUTPUT_DIR}/<store_id>.xlsx에 저장)")
    print(f"수집 기준: 최근 6개월 (기준일 {CUTOFF} 이후)")

    all_rows = []
    start_time = time.time()

    for i, store in enumerate(store_list, 1):
        if i > 1:
            await asyncio.sleep(random.uniform(*INTER_STORE_DELAY_RANGE_SEC))

        store_start = time.time()
        store_path = output_path_for(store["store_id"])
        seen_keys, next_seq = load_existing_reviews(store_path)
        if seen_keys:
            print(f"  기존 {store_path}에서 리뷰 {len(seen_keys)}개 확인. 이미 수집된 리뷰는 건너뛰고 새 리뷰만 추가합니다.")

        print(f"\n[{i}/{limit}] {store['name']} ({store['naver_id']}) 리뷰 수집 중...")

        seq_holder = [next_seq.get(store["store_id"], 0)]
        store_rows: list = []

        async def on_pass(new_reviews, store=store, seq_holder=seq_holder, store_rows=store_rows, store_path=store_path):
            pass_rows = []
            for r in new_reviews:
                seq_holder[0] += 1
                pass_rows.append(review_to_row(store, r, seq_holder[0]))
            store_rows.extend(pass_rows)
            save_to_excel(pass_rows, store_path)  # persist immediately -- a store can need
            # many browser restarts, so don't wait until it's fully done to save

        try:
            reviews = await get_store_reviews(store["naver_id"], seen_keys, store["store_id"], on_pass=on_pass)
        except Exception as e:
            print(f"  실패: {e}")
            reviews = []

        all_rows.extend(store_rows)

        store_elapsed = time.time() - store_start
        total_elapsed = time.time() - start_time
        avg_per_store = total_elapsed / i
        remaining = (limit - i) * avg_per_store

        print(f"  -> 신규 {len(reviews)}개 리뷰 수집 및 저장 완료 (이 매장 {store_elapsed:.1f}초 소요)")
        print(f"  진행률: {i}/{limit} ({i/limit*100:.1f}%) | "
              f"누적 신규 리뷰 {len(all_rows)}개 | "
              f"경과 {format_eta(total_elapsed)} | "
              f"예상 남은 시간 {format_eta(remaining)}")

    total_elapsed = time.time() - start_time
    print(f"\n완료: 신규 {len(all_rows)}개 리뷰 저장 -> {OUTPUT_DIR}/ (총 소요시간 {format_eta(total_elapsed)})")


if __name__ == "__main__":
    asyncio.run(main())

"""One-time migration: splits the legacy single Review.xlsx / Review_category.xlsx
into per-store files (reviews_by_store/<store_id>.xlsx, tags_by_store/<store_id>.xlsx).

Run once, then Review.xlsx / Review_category.xlsx are no longer read by anything
(Review.py, Review_category.py, refresh_store.py, load_data.py all use the
per-store folders now) and can be archived or deleted.

Usage:
    python split_review_files.py
"""
import os
from collections import defaultdict

from openpyxl import Workbook, load_workbook

BASE = os.path.dirname(__file__)


def split(input_path, output_dir, store_id_column):
    if not os.path.exists(input_path):
        print(f"{input_path} 없음, 건너뜀")
        return

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = list(next(it))
    store_idx = headers.index(store_id_column)

    rows_by_store = defaultdict(list)
    for row in it:
        if row[0] is None:
            continue
        rows_by_store[row[store_idx]].append(row)

    os.makedirs(output_dir, exist_ok=True)
    for store_id, store_rows in rows_by_store.items():
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.append(headers)
        for row in store_rows:
            out_ws.append(list(row))
        out_wb.save(os.path.join(output_dir, f"{store_id}.xlsx"))

    total_rows = sum(len(v) for v in rows_by_store.values())
    print(f"{input_path}: {total_rows}건 -> {len(rows_by_store)}개 매장 파일로 분리 ({output_dir}/)")


def main():
    split(os.path.join(BASE, "Review.xlsx"), os.path.join(BASE, "reviews_by_store"), "store_id")
    split(os.path.join(BASE, "Review_category.xlsx"), os.path.join(BASE, "tags_by_store"), "store_id")
    print("완료.")


if __name__ == "__main__":
    main()

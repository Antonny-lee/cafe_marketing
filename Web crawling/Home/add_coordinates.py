"""Adds lat/lng columns to the existing cafes.xlsx by re-fetching each store's
place-detail page just for its coordinate. Naver Map already returns this in
PlaceDetailBase.coordinate, so no external geocoding API is needed."""
import time

from openpyxl import load_workbook

from home import get_place_detail_row

SRC = "cafes.xlsx"
REQUEST_DELAY_SEC = 0.6


def main():
    wb = load_workbook(SRC)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "lat" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="lat")
        ws.cell(row=1, column=len(headers) + 2, value="lng")
        lat_col, lng_col = len(headers) + 1, len(headers) + 2
    else:
        lat_col, lng_col = headers.index("lat") + 1, headers.index("lng") + 1

    naver_id_col = headers.index("naver_id") + 1
    rows = list(ws.iter_rows(min_row=2))
    print(f"{len(rows)}개 매장 좌표 갱신 중...")

    for i, row in enumerate(rows, 1):
        pid = str(row[naver_id_col - 1].value)
        store_id = row[0].value
        try:
            detail = get_place_detail_row(pid)
        except Exception as e:
            print(f"[{i}/{len(rows)}] {store_id} 실패: {e}")
            continue
        if detail:
            ws.cell(row=row[0].row, column=lat_col, value=detail.get("lat"))
            ws.cell(row=row[0].row, column=lng_col, value=detail.get("lng"))
            print(f"[{i}/{len(rows)}] {store_id} {detail.get('lat')},{detail.get('lng')}")
        time.sleep(REQUEST_DELAY_SEC)

    wb.save(SRC)
    print("완료.")


if __name__ == "__main__":
    main()

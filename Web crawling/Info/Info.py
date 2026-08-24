import json
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from scrapling.fetchers import Fetcher

CAFES_XLSX = "../Home/cafes.xlsx"
OUTPUT_XLSX = "Info.xlsx"
DETAIL_URL_TMPL = "https://m.place.naver.com/place/{pid}/information"
REQUEST_DELAY_SEC = 0.6

# Fixed order for the sheets we already know about (matches the real page's
# section order). Any section discovered later that isn't in here just gets
# appended as a new sheet -- see `sheets` dict in main().
KNOWN_SHEET_ORDER = ["소개", "편의시설 및 서비스", "노키즈존", "반려동물 동반",
                     "주차", "좌석.공간", "결제수단", "SNS"]


def extract_apollo_state(html: str):
    marker = "window.__APOLLO_STATE__ = "
    idx = html.find(marker)
    if idx == -1:
        return None
    start = idx + len(marker)
    i = start
    depth = 0
    in_str = False
    esc = False
    while i < len(html):
        c = html[i]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
        else:
            if c == '"':
                in_str = True
            elif c == "{":
                depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    i += 1
                    break
        i += 1
    return json.loads(html[start:i])


def deref(apollo: dict, obj):
    if isinstance(obj, dict) and "__ref" in obj:
        return apollo.get(obj["__ref"])
    return obj


def load_stores(path: str):
    wb = load_workbook(path)
    ws = wb.active
    stores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        store_id, name, _loc, _subway, _hours, naver_id = row[:6]
        if not naver_id:
            continue
        stores.append({"store_id": store_id, "name": name, "naver_id": str(naver_id)})
    return stores


def find_place_detail(apollo: dict, pid: str):
    root = apollo.get("ROOT_QUERY", {})
    for k, v in root.items():
        if k.startswith("placeDetail(") and f'"id":"{pid}"' in k:
            return v
    for k, v in root.items():
        if k.startswith("placeDetail("):
            return v
    return None


def get_store_sections(store: dict):
    """Returns dict[sheet_name] -> list of row dicts for this store."""
    url = DETAIL_URL_TMPL.format(pid=store["naver_id"])
    page = Fetcher.get(url, stealthy_headers=True)
    html = page.body.decode("utf-8")
    apollo = extract_apollo_state(html)
    if apollo is None:
        return {}

    base = apollo.get(f"PlaceDetailBase:{store['naver_id']}")
    place_detail = find_place_detail(apollo, store["naver_id"])
    if base is None or place_detail is None:
        return {}

    def row(**kw):
        return {"store_id": store["store_id"], "store_name_ref": store["name"], **kw}

    sections: dict = {}

    # 소개
    for k, v in place_detail.items():
        if k.startswith("description(") and v:
            sections.setdefault("소개", []).append(row(intro_text=v))
            break

    info_tab = None
    for k, v in place_detail.items():
        if k.startswith("informationTab("):
            info_tab = v
            break

    if info_tab:
        # 편의시설 및 서비스
        for ref in info_tab.get("facilities") or []:
            fac = deref(apollo, ref)
            if fac and fac.get("name"):
                sections.setdefault("편의시설 및 서비스", []).append(
                    row(item_text=fac["name"], detail=""))

        # 노키즈존
        no_kids = info_tab.get("noKidsZone")
        if no_kids and no_kids.get("isOperated"):
            sections.setdefault("노키즈존", []).append(
                row(item_text="노키즈존", detail=no_kids.get("description") or ""))

        # 반려동물 동반
        pet = info_tab.get("pet")
        if pet is not None:
            allowed = pet.get("isAllowed")
            desc = pet.get("description")
            fallback = "가능" if allowed else "불가"
            sections.setdefault("반려동물 동반", []).append(
                row(item_text="반려동물 동반", detail=desc or fallback))

        # 주차
        parking = info_tab.get("parkingInfo")
        if parking is not None:
            desc = parking.get("description")
            if not desc and parking.get("basicParking"):
                free = parking["basicParking"].get("isFree")
                desc = "무료" if free else "유료"
            sections.setdefault("주차", []).append(
                row(item_text="주차", detail=desc or ""))

    if "편의시설 및 서비스" not in sections and base.get("conveniences"):
        # fall back to the short convenience-tag list when the richer
        # informationTab facilities aren't available for this store
        for c in base["conveniences"]:
            sections.setdefault("편의시설 및 서비스", []).append(row(item_text=c, detail=""))

    # 좌석.공간
    tab_details = place_detail.get("restaurantInfoTabDetails")
    if tab_details:
        for ref in tab_details.get("seat") or []:
            seat = deref(apollo, ref)
            if seat and seat.get("name"):
                sections.setdefault("좌석.공간", []).append(
                    row(item_text=seat["name"], detail=seat.get("description") or ""))

    # 결제수단
    for method in base.get("paymentInfo") or []:
        sections.setdefault("결제수단", []).append(row(item_text=method, detail=""))

    # SNS
    homepages = place_detail.get("homepages")
    if homepages:
        links = []
        if homepages.get("repr"):
            links.append(homepages["repr"])
        links.extend(homepages.get("etc") or [])
        for link in links:
            sections.setdefault("SNS", []).append(
                row(item_text=link.get("type", ""), detail=link.get("url", "")))

    return sections


def save_to_excel(all_sections: dict, path: str):
    wb = Workbook()
    wb.remove(wb.active)

    ordered_names = [n for n in KNOWN_SHEET_ORDER if n in all_sections]
    ordered_names += [n for n in all_sections if n not in ordered_names]

    wrap = Alignment(wrap_text=True, vertical="top")
    for name in ordered_names:
        rows = all_sections[name]
        ws = wb.create_sheet(name[:31])  # Excel sheet name limit
        if name == "소개":
            headers = ["store_id", "store_name_ref", "intro_text"]
            widths = [10, 22, 70]
        else:
            headers = ["store_id", "store_name_ref", "item_text", "detail"]
            widths = [10, 22, 24, 40]

        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        for col_letter, width in zip("ABCDEFGH", widths):
            ws.column_dimensions[col_letter].width = width
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                cell.alignment = wrap

    for attempt in range(1, 6):
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"  {path} 파일이 열려 있어 저장할 수 없습니다. 닫아주세요. ({attempt}/5, 5초 후 재시도)")
            time.sleep(5)
    wb.save(path)


def main():
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 5

    print("cafes.xlsx에서 매장 목록 읽는 중...")
    stores = load_stores(CAFES_XLSX)
    print(f"총 {len(stores)}개 매장. {limit}개 처리합니다.")

    all_sections: dict = {}
    for i, store in enumerate(stores[:limit], 1):
        print(f"[{i}/{limit}] {store['name']} ({store['naver_id']}) 정보 가져오는 중...")
        try:
            sections = get_store_sections(store)
        except Exception as e:
            print(f"  실패: {e}")
            sections = {}

        for name, rows in sections.items():
            all_sections.setdefault(name, []).extend(rows)

        summary = ", ".join(f"{k} {len(v)}" for k, v in sections.items())
        print(f"  -> {summary if summary else '데이터 없음'}")
        time.sleep(REQUEST_DELAY_SEC)

    new_sheets = [n for n in all_sections if n not in KNOWN_SHEET_ORDER]
    if new_sheets:
        print(f"\n새로운 섹션 발견: {new_sheets}")

    save_to_excel(all_sections, OUTPUT_XLSX)
    total = sum(len(v) for v in all_sections.values())
    print(f"완료: {len(all_sections)}개 시트, 총 {total}개 행 -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()

import asyncio
import json
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from scrapling.fetchers import AsyncStealthySession, Fetcher
from scrapling.parser import Selector

CAFES_XLSX = "../Home/cafes.xlsx"
OUTPUT_XLSX = "AIBriefing.xlsx"
HOME_URL_TMPL = "https://pcmap.place.naver.com/restaurant/{pid}/home"
STATIC_HOME_URL_TMPL = "https://m.place.naver.com/place/{pid}/home"
REQUEST_DELAY_SEC = 0.5


def extract_apollo_state(html: str):
    marker = "window.__APOLLO_STATE__ = "
    idx = html.find(marker)
    if idx == -1:
        return None
    start = idx + len(marker)
    i = start
    depth = 0
    in_str = False
    esc = False
    while i < len(html):
        c = html[i]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
        else:
            if c == '"':
                in_str = True
            elif c == "{":
                depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    i += 1
                    break
        i += 1
    return json.loads(html[start:i])


def find_place_detail(apollo: dict, pid: str):
    root = apollo.get("ROOT_QUERY", {})
    for k, v in root.items():
        if k.startswith("placeDetail(") and f'"id":"{pid}"' in k:
            return v
    for k, v in root.items():
        if k.startswith("placeDetail("):
            return v
    return None


def load_stores(path: str):
    wb = load_workbook(path)
    ws = wb.active
    stores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        store_id, name, _loc, _subway, _hours, naver_id = row[:6]
        if not naver_id:
            continue
        stores.append({"store_id": store_id, "name": name, "naver_id": str(naver_id)})
    return stores


def has_ai_briefing(store: dict) -> bool:
    """Cheap static check so we only pay for a browser launch on stores that
    actually have the AI briefing widget."""
    url = STATIC_HOME_URL_TMPL.format(pid=store["naver_id"])
    page = Fetcher.get(url, stealthy_headers=True)
    apollo = extract_apollo_state(page.body.decode("utf-8"))
    if apollo is None:
        return False
    place_detail = find_place_detail(apollo, store["naver_id"])
    return bool(place_detail and place_detail.get("hasAiBriefing"))


UNIT_CHARS = "층명개년월일시분원잔개월살번"


def clean_sentence(text: str) -> str:
    text = re.sub(r"\s+([,.)])", r"\1", text)
    text = re.sub(r"([(])\s+", r"\1", text)
    text = re.sub(r"\s+([(])", r" \1", text)  # keep exactly one space before "("
    text = re.sub(r"([(])\s+", r"\1", text)   # ...but none right after it
    text = re.sub(r"[ \t]+([)])", r"\1", text)
    # a short trailing particle right after ")" (e.g. "등) 를" -> "등)를")
    text = re.sub(r"([)])\s+(을|를|이|가|은|는|의|에|와|과|도|만)(?=[\s.,]|$)", r"\1\2", text)
    # Naver sometimes tokenizes a number and its immediately following unit as
    # separate spans (e.g. "2 층"); collapse that specific number+unit gap.
    text = re.sub(rf"(\d)\s+(?=[{UNIT_CHARS}])", r"\1", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


async def get_ai_briefing_sentences(pid: str):
    url = HOME_URL_TMPL.format(pid=pid)
    sentences: list = []

    async def action(page):
        await page.wait_for_timeout(1200)
        found = False
        for _ in range(15):
            if await page.get_by_text("AI 브리핑").count() > 0:
                found = True
                break
            await page.mouse.wheel(0, 800)
            await page.wait_for_timeout(400)
        if not found:
            return page

        await page.wait_for_timeout(800)
        html = await page.content()
        sel = Selector(html)
        for li in sel.css("ul.pAVCt li.cqLJy"):
            span = li.css("div.XOgRw > span.qGckC")
            if span:
                text = span[0].get_all_text(strip=True, separator=" ")
                sentences.append(clean_sentence(text))
        return page

    async with AsyncStealthySession(headless=True, disable_resources=True) as session:
        await session.fetch(url, page_action=action, network_idle=True, timeout=60000)

    return sentences


HEADERS = ["store_id", "store_name_ref", "sentence"]


def save_to_excel(rows: list[dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "AI브리핑"
    ws.append(HEADERS)

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])

    for col_letter, width in zip("ABC", [10, 22, 70]):
        ws.column_dimensions[col_letter].width = width
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.alignment = wrap

    for attempt in range(1, 6):
        try:
            wb.save(path)
            return
        except PermissionError:
            print(f"  {path} 파일이 열려 있어 저장할 수 없습니다. 닫아주세요. ({attempt}/5, 5초 후 재시도)")
            time.sleep(5)
    wb.save(path)


async def main():
    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:
        pass

    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 5

    print("cafes.xlsx에서 매장 목록 읽는 중...")
    stores = load_stores(CAFES_XLSX)
    print(f"총 {len(stores)}개 매장. {limit}개 처리합니다.")

    all_rows = []
    for i, store in enumerate(stores[:limit], 1):
        print(f"[{i}/{limit}] {store['name']} ({store['naver_id']}) 확인 중...")
        try:
            if not has_ai_briefing(store):
                print("  -> AI 브리핑 없음, 건너뜀")
                time.sleep(REQUEST_DELAY_SEC)
                continue
        except Exception as e:
            print(f"  확인 실패: {e}")
            continue

        try:
            sentences = await get_ai_briefing_sentences(store["naver_id"])
        except Exception as e:
            print(f"  수집 실패: {e}")
            sentences = []

        print(f"  -> {len(sentences)}개 문장 수집")
        for s in sentences:
            all_rows.append({"store_id": store["store_id"], "store_name_ref": store["name"], "sentence": s})

    save_to_excel(all_rows, OUTPUT_XLSX)
    print(f"완료: {len(all_rows)}개 문장 저장 -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    asyncio.run(main())
